from datetime import datetime
import os
import logging
from typing import Optional

from django.core.management.base import BaseCommand, CommandError
from slugify import slugify
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import exceptions as openpyxl_exceptions

from saleor.attribute.models.base import AttributeValue
from saleor.attribute.models.product import AssignedProductAttributeValue
from saleor.ornament.vendors.attribute_utils import AttributeUtils
from saleor.ornament.vendors.utils import (
    form_description,
)
from saleor.product.models import (
    Product,
    ProductTranslation,
    ProductChannelListing,
)
from saleor.channel.models import Channel


logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = "Import/Update Checkup products with XLSX file"
    known_header_rows = [
        "sku",
        "title ru",
        "title en",
        "title de",
        "title pt",
        "title es",
        "description ru",
        "description en",
        "description de",
        "description pt",
        "description es",
        "biomarkers",
        "medical_exams",
    ]
    sku_missed_en_translation = []
    sku_missed_medical_data = []
    checkup_category_id = 1
    product_category_id = 1
    checkup_default_language_code = "en"
    default_channel_slug = "uk-gettested"
    default_channel_currency = "GBP"

    def add_arguments(self, parser):
        parser.add_argument("filename", help="Path to XLSX file with target data")
        parser.add_argument(
            "--channel_slug",
            default=self.default_channel_slug,
            help=f'Specifies the channel where added/updated products should be listed to use. Default is "{self.default_channel_slug}".',
        )

    def check_sheet_header_cols(self, sheet: Worksheet) -> None:
        sheet_header_cols: set = set(
            [s for s in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0]
        )

        if set(self.known_header_rows).difference(sheet_header_cols):
            raise CommandError(
                f"Sheet header does not contain all known columns: {self.known_header_rows}"
            )

    def convert_medical_data_ids(self, data_row: Optional[str]) -> Optional[list[int]]:
        try:
            if data_row and isinstance(data_row, str):
                return list(map(int, data_row.split(",")))
            else:
                return None
        except ValueError:
            return None

    def check_row_medical_data(self, row: tuple) -> bool:
        biomarkers = self.convert_medical_data_ids(row[11])
        medical_exams = self.convert_medical_data_ids(row[12])

        if not biomarkers and not medical_exams:
            self.sku_missed_medical_data.append(row[0])

        return bool(biomarkers) or bool(medical_exams)

    def check_row_required_data(self, row: tuple) -> bool:
        if row[0] and not row[2]:
            self.sku_missed_en_translation.append(row[0])
        return row[2] and self.check_row_medical_data(row)

    def get_medical_attribute_values_ids(self, attribute_name: str) -> dict[int, int]:
        medical_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids[attribute_name]
        ).values_list("pk", "name")

        return {int(b[1]): b[0] for b in medical_attribute_values}

    def handle(self, *args, **options):
        filename = options.get("filename")
        channel_slug = options.get("channel_slug")

        if not filename or not os.path.isfile(filename):
            raise CommandError(f'Source file "{filename}" does not exist.')

        try:
            wb = load_workbook(filename)
        except openpyxl_exceptions.InvalidFileException:
            raise CommandError(
                f"openpyxl does not support the old .xls file format, please use .xlsx file format."
            )

        sheet = wb["translations"]

        self.check_sheet_header_cols(sheet)

        data = {
            s[0]: {
                "sku": s[0],
                "title": {
                    "ru": s[1],
                    "en": s[2],
                    "de": s[3],
                    "pt": s[4],
                    "es": s[5],
                },
                "description": {
                    "ru": s[6],
                    "en": s[7],
                    "de": s[8],
                    "pt": s[9],
                    "es": s[10],
                },
                "biomarkers": self.convert_medical_data_ids(str(s[11])) or [],
                "medical_exams": self.convert_medical_data_ids(str(s[12])) or [],
            }
            for s in sheet.iter_rows(min_row=2, values_only=True)
            if self.check_row_required_data(s)
        }

        if self.sku_missed_en_translation:
            logger.info(
                f"No EN `title` translation for SKUs {self.sku_missed_en_translation}"
            )

        if self.sku_missed_medical_data:
            logger.info(
                f"No biomarkers or medical_exams for SKUs {self.sku_missed_medical_data}"
            )

        current_products = Product.objects.values_list("name", flat=True)

        data_to_insert = {k: v for k, v in data.items() if k not in current_products}
        data_to_update = {k: v for k, v in data.items() if k in current_products}

        channel = Channel.objects.filter(slug=channel_slug).first()

        if not channel:
            raise CommandError(f"Can't find a channel with slug `{channel_slug}`!")

        new_products = []

        assigned_product_attribute_values_to_insert: list[
            AssignedProductAttributeValue
        ] = []
        assigned_product_attribute_values_to_delete: list[
            AssignedProductAttributeValue
        ] = []

        biomarkers_attribute_values_ids = self.get_medical_attribute_values_ids(
            "biomarkers"
        )
        medical_exams_attribute_values_ids = self.get_medical_attribute_values_ids(
            "medical_exams"
        )

        for d in data_to_insert.values():
            title = d["title"][self.checkup_default_language_code]
            description = d["description"][self.checkup_default_language_code]

            new_products.append(
                Product(
                    product_type_id=self.product_category_id,
                    name=d["sku"],
                    slug=slugify(d["sku"], lowercase=True, max_length=100),
                    description=form_description(title, description),
                    description_plaintext=description or "",
                    category_id=self.checkup_category_id,
                    search_index_dirty=True,
                )
            )

        inserted_products = Product.objects.bulk_create(
            new_products, ignore_conflicts=False
        )

        for product in inserted_products:
            product_data = data_to_insert.get(product.name) or {}
            biomarkers = product_data.get("biomarkers")
            medical_exams = product_data.get("medical_exams")

            if biomarkers:
                assigned_product_attribute_values_to_insert += (
                    AttributeUtils.add_medical_attributes_data(
                        product.pk,
                        biomarkers,
                        biomarkers_attribute_values_ids,
                    )
                )

            if medical_exams:
                assigned_product_attribute_values_to_insert += (
                    AttributeUtils.add_medical_attributes_data(
                        product.pk,
                        medical_exams,
                        medical_exams_attribute_values_ids,
                    )
                )

        products_to_update = Product.objects.filter(name__in=data_to_update.keys())

        for product in products_to_update:
            data_product = data_to_update.get(product.name)

            if data_product:
                biomarkers = data_product.get("biomarkers")
                medical_exams = data_product.get("medical_exams")

                if biomarkers:
                    biomarkers = [int(b) for b in biomarkers]

                    biomarkers_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "biomarkers"
                            ],
                            product_id=product.pk,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        biomarkers_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            biomarkers,
                            biomarkers_attribute_values_ids,
                        )
                    )

                if medical_exams:
                    medical_exams = [int(b) for b in medical_exams]

                    medical_exams_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "medical_exams"
                            ],
                            product_id=product.pk,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        medical_exams_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            medical_exams,
                            medical_exams_attribute_values_ids,
                        )
                    )

        all_checkup_products = Product.objects.filter(
            name__in=data.keys()
        ).prefetch_related("translations")

        now = datetime.now()

        product_channel_listings = []
        product_translations_to_insert = []
        product_translations_to_update = []

        for product in all_checkup_products:
            product_channel_listings.append(
                ProductChannelListing(
                    channel=channel,
                    product=product,
                    visible_in_listings=True,
                    currency=channel.currency_code,
                    is_published=True,
                    published_at=now,
                )
            )

            data_product = data.get(product.name)

            if not data_product:
                continue

            current_translations: list[ProductTranslation] = [
                t for t in product.translations.all()
            ]

            if current_translations:
                for translation in current_translations:
                    title = data_product["title"].pop(translation.language_code, None)
                    description = data_product["description"].pop(
                        translation.language_code, None
                    )

                    if title:
                        translation.name = product.name
                        translation.description = form_description(title, description)
                        product_translations_to_update.append(translation)

            for lang, value in data_product["title"].items():
                if value:
                    product_translations_to_insert.append(
                        ProductTranslation(
                            product=product,
                            name=product.name,
                            description=form_description(
                                value, data_product["description"].get(lang)
                            ),
                            language_code=lang,
                        )
                    )

        AssignedProductAttributeValue.objects.filter(
            id__in=[apav.pk for apav in assigned_product_attribute_values_to_delete]
        ).delete()

        AssignedProductAttributeValue.objects.bulk_create(
            assigned_product_attribute_values_to_insert,
            batch_size=500,
        )

        ProductChannelListing.objects.bulk_create(
            product_channel_listings, ignore_conflicts=True
        )
        ProductTranslation.objects.bulk_create(product_translations_to_insert)
        ProductTranslation.objects.bulk_update(
            product_translations_to_update, ["name", "description"]
        )

        return
