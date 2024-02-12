from datetime import datetime
import enum
import os
import re
from itertools import chain
from typing import Optional

from django.core.management.base import BaseCommand, CommandError
from slugify import slugify
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import exceptions as openpyxl_exceptions

from saleor.attribute.models.base import AttributeValue
from saleor.attribute.models.product import AssignedProductAttributeValue
from saleor.ornament.vendors.attribute_utils import AttributeUtils
from saleor.ornament.vendors.utils import form_description, form_rich_text
from saleor.product.models import (
    Product,
    ProductVariant,
    ProductChannelListing,
    CollectionProduct,
    Category,
    Collection,
    ProductVariantChannelListing,
)
from saleor.channel.models import Channel


class KdlDurationUnitEnum(enum.Enum):
    HOUR = 1
    DAY = 2


class Command(BaseCommand):
    help = "Import/Update KDL products with XLSX file"
    known_header_rows = [
        "SKU",
        "Наименование теста (KDL)",
        "Категория",
        "Подготовка к исследованию",
        "Описание",
        "Срок",
        "Биоматериал",
        "Биомаркеры",
        "medical_exam_type_object",
        "is popular",
        "is hidden",
        "rating",
        "color slug",
    ]

    def add_arguments(self, parser):
        parser.add_argument("filename", help="Path to XLSX file with target data")

    def check_kdl_sheet_header_cols(self, sheet: Worksheet) -> None:
        sheet_header_cols: set = set(
            [s for s in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0]
        )

        if set(self.known_header_rows).difference(sheet_header_cols):
            raise CommandError(
                f"Sheet header does not contain all known columns: {self.known_header_rows}"
            )

    def get_medical_data_ids(self, col_data: str, ids: list[int]) -> list[int]:
        data_strings = col_data.split("\n")

        try:
            found_ids = chain.from_iterable(
                [re.findall(r"\d+", b.split("-")[-1]) for b in data_strings if b]
            )
            return [id for id in found_ids if int(id) in ids]
        except (IndexError, ValueError):
            raise CommandError(f"Can not parse medical data column: {col_data}")

    def add_featured_attribute_data(
        self,
        product_id: int,
        featured_attribute_values: dict[str, AttributeValue],
    ) -> AssignedProductAttributeValue:
        featured_slug = f"{AttributeUtils.attrubutes_ids['featured']}_true"

        featured_attribute_value = featured_attribute_values.get(featured_slug)

        if not featured_attribute_value:
            raise CommandError(
                f"""Can't find featured attribute value for slug {featured_slug}"""
            )

        return AssignedProductAttributeValue(
            value_id=featured_attribute_value.pk,
            product_id=product_id,
        )

    def check_row_required_data(self, row: tuple) -> bool:
        return all([row[0], row[1], row[2]])

    def get_attribute_value(
        self,
        product_id: int,
        attribute_name: str,
        known_values: dict[str, AttributeValue],
    ) -> Optional[AttributeValue]:
        slug = f"{product_id}_{AttributeUtils.attrubutes_ids[attribute_name]}"
        return known_values.get(slug)

    def should_recreate_attribute_assignment(
        self,
        product: Product,
        db_attribute_value: AttributeValue,
    ) -> bool:
        assigned_value = product.attributevalues.filter(
            value_id=db_attribute_value.pk
        ).first()

        return not assigned_value

    def handle(self, *args, **options):
        filename = options.get("filename")

        if not filename or not os.path.isfile(filename):
            raise CommandError(f'Source file "{filename}" does not exist.')

        try:
            wb = load_workbook(filename)
        except openpyxl_exceptions.InvalidFileException:
            raise CommandError(
                f"openpyxl does not support the old .xls file format, please use .xlsx file format."
            )

        sheet = wb["KDL-import"]

        self.check_kdl_sheet_header_cols(sheet)

        medical_data = fetch_medical_data()

        data = {
            s[0]: {
                "sku": s[0],
                "name": s[1],
                "category": s[2],
                "preparation": s[3],
                "description": s[4],
                "duration": s[5],
                "biomaterial": s[6],
                "biomarkers": self.get_medical_data_ids(
                    str(s[7]), medical_data.biomarker_ids
                ),
                "medical_exams": self.get_medical_data_ids(
                    str(s[8]), medical_data.medical_exams_ids
                ),
                "is_popular": s[9],
                "is_hidden": s[10],
                "rating": s[11],
                "color_slug": s[12],
            }
            for s in sheet.iter_rows(min_row=2, values_only=True)
            if self.check_row_required_data(s)
        }

        current_products = Product.objects.values_list("name", flat=True)

        categories = Category.objects.all()
        categories = {c.name: c for c in categories}

        data_to_update = {k: v for k, v in data.items() if k in current_products}
        data_to_insert = {
            k: v
            for k, v in data.items()
            if k not in current_products and v["is_hidden"] != 1
        }

        channels = Channel.objects.filter(is_active=True)
        popular_collection = Collection.objects.filter(name="Популярное").first()

        biomaterial_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["kdl-biomaterials"]
        )
        biomaterial_attribute_values = {v.slug: v for v in biomaterial_attribute_values}

        preparation_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["kdl-preparation"]
        )
        preparation_attribute_values = {v.slug: v for v in preparation_attribute_values}

        duration_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["kdl-max_duration"]
        )
        duration_attribute_values = {v.slug: v for v in duration_attribute_values}

        duration_unit_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["kdl-duration_unit"]
        )
        duration_unit_attribute_values = {
            v.slug: v for v in duration_unit_attribute_values
        }

        featured_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["featured"]
        )
        featured_attribute_values = {v.slug: v for v in featured_attribute_values}

        color_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["color"]
        )
        color_attribute_values = {v.slug: v for v in color_attribute_values}

        popular_collection_assignments = CollectionProduct.objects.filter(
            collection=popular_collection
        )
        popular_collection_assignments = {
            v.product.pk: v for v in popular_collection_assignments
        }

        new_products = []

        for d in data_to_insert.values():
            category = categories.get(d["category"])

            if not category:
                raise CommandError(
                    f"""Can't find category: {d["category"]}, product: {d["sku"]}"""
                )

            new_products.append(
                Product(
                    product_type_id=1,
                    name=d["sku"],
                    slug=slugify(d["sku"], lowercase=True, max_length=100),
                    description=form_description(d["name"], d["description"]),
                    description_plaintext=d["description"] or "",
                    category_id=category.pk,
                    search_index_dirty=True,
                    rating=d["rating"],
                )
            )

        inserted_products = Product.objects.bulk_create(
            new_products, ignore_conflicts=False
        )

        product_variants = []
        product_channel_listings = []
        assigned_product_attribute_values = []
        collection_products = []

        now = datetime.now()

        biomarkers_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["biomarkers"]
        ).values_list("pk", "name")
        biomarkers_attribute_values_ids: dict[int, int] = {
            int(b[1]): b[0] for b in biomarkers_attribute_values
        }
        medical_exams_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["medical_exams"]
        ).values_list("pk", "name")
        medical_exams_attribute_values_ids: dict[int, int] = {
            int(m[1]): m[0] for m in medical_exams_attribute_values
        }

        for product in inserted_products:
            data_product = data.get(product.name)

            if data_product:
                product_variants.append(
                    ProductVariant(
                        sku=f'KDL-{data_product["sku"]}',
                        name="KDL",
                        product=product,
                        track_inventory=False,
                    )
                )

                for channel in channels:
                    product_channel_listings.append(
                        ProductChannelListing(
                            channel=channel,
                            product=product,
                            visible_in_listings=True,
                            available_for_purchase_at=now,
                            currency="RUB",
                            is_published=True,
                            published_at=now,
                        )
                    )

                if (
                    data_product["biomaterial"]
                    and "N/A" not in data_product["biomaterial"]
                ):
                    assigned_product_attribute_values.append(
                        AttributeUtils.add_biomaterial_attribute_data(
                            product.pk,
                            data_product["biomaterial"],
                        )
                    )

                if (
                    data_product["preparation"]
                    and "N/A" not in data_product["preparation"]
                ):
                    assigned_product_attribute_values.append(
                        AttributeUtils.add_preparation_attribute_data(
                            product.pk,
                            data_product["preparation"],
                        )
                    )

                if data_product["duration"] and "N/A" not in str(
                    data_product["duration"]
                ):
                    assigned_product_attribute_values.append(
                        AttributeUtils.add_numeric_attribute_data(
                            product.pk,
                            int(data_product["duration"]),
                            AttributeUtils.attrubutes_ids["kdl-max_duration"],
                        )
                    )
                    assigned_product_attribute_values.append(
                        AttributeUtils.add_numeric_attribute_data(
                            product.pk,
                            # 2 - legacy: duration unit DAYS
                            KdlDurationUnitEnum.DAY.value,
                            AttributeUtils.attrubutes_ids["kdl-duration_unit"],
                        )
                    )

                if data_product["biomarkers"]:
                    assigned_product_attribute_values += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            [int(b) for b in data_product["biomarkers"]],
                            biomarkers_attribute_values_ids,
                        )
                    )

                if data_product["medical_exams"]:
                    assigned_product_attribute_values += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            [int(m) for m in data_product["medical_exams"]],
                            medical_exams_attribute_values_ids,
                        )
                    )

                if data_product["is_popular"] == 1:
                    assigned_product_attribute_values.append(
                        self.add_featured_attribute_data(
                            product.pk, featured_attribute_values
                        )
                    )

                    if popular_collection:
                        collection_products.append(
                            CollectionProduct(
                                collection_id=popular_collection.pk,
                                product_id=product.pk,
                            )
                        )

                if data_product["color_slug"]:
                    assigned_product_attribute_values.append(
                        AttributeUtils.add_color_attribute_data(
                            product.pk,
                            data_product["color_slug"],
                            color_attribute_values,
                        )
                    )

        ProductChannelListing.objects.bulk_create(product_channel_listings)
        ProductVariant.objects.bulk_create(product_variants)

        CollectionProduct.objects.bulk_create(collection_products)
        AssignedProductAttributeValue.objects.bulk_create(
            assigned_product_attribute_values
        )

        # UPDATE
        data_to_delist = [
            f"KDL-{k}" for k, v in data_to_update.items() if v["is_hidden"] == 1
        ]
        data_to_update = {
            k: v for k, v in data_to_update.items() if v["is_hidden"] != 1
        }

        products_to_update = Product.objects.filter(
            name__in=data_to_update.keys()
        ).prefetch_related("attributevalues")

        attribute_values_to_update: list[AttributeValue] = []
        assigned_product_attribute_values_to_insert: list[
            AssignedProductAttributeValue
        ] = []
        attribute_values_to_delete: list[AttributeValue] = []
        assigned_product_attribute_values_to_delete: list[
            AssignedProductAttributeValue
        ] = []
        collection_products_to_insert = []

        for p in products_to_update:
            data_product = data_to_update.get(p.name)
            product_id = p.pk

            if data_product:
                category = categories.get(data_product["category"])
                if category:
                    p.category = category

                p.description = form_description(
                    data_product["name"], data_product["description"]
                )
                p.description_plaintext = data_product["description"] or ""

                if data_product["rating"]:
                    p.rating = float(data_product["rating"])
                else:
                    p.rating = None

                p.save()

                # BIOMATERIAL

                biomaterial_db_attribute_value = self.get_attribute_value(
                    product_id, "kdl-biomaterials", biomaterial_attribute_values
                )

                if (
                    data_product["biomaterial"]
                    and "N/A" not in data_product["biomaterial"]
                ):
                    insert = False

                    name = data_product["biomaterial"].replace("\n", ", ")

                    if biomaterial_db_attribute_value:
                        recreate = self.should_recreate_attribute_assignment(
                            p,
                            biomaterial_db_attribute_value,
                        )

                        if recreate:
                            attribute_values_to_delete.append(
                                biomaterial_db_attribute_value
                            )
                            insert = True
                        else:
                            biomaterial_db_attribute_value.name = name
                            biomaterial_db_attribute_value.plain_text = name
                            attribute_values_to_update.append(
                                biomaterial_db_attribute_value
                            )
                    else:
                        insert = True

                    if insert:
                        assigned_product_attribute_values_to_insert.append(
                            AttributeUtils.add_biomaterial_attribute_data(
                                product_id,
                                data_product["biomaterial"],
                            )
                        )
                else:
                    if biomaterial_db_attribute_value:
                        attribute_values_to_delete.append(
                            biomaterial_db_attribute_value
                        )

                # PREPARATION

                preparation_db_attribute_value = self.get_attribute_value(
                    product_id, "kdl-preparation", preparation_attribute_values
                )

                if (
                    data_product["preparation"]
                    and "N/A" not in data_product["preparation"]
                ):
                    insert = False

                    preparation = data_product["preparation"]
                    name = preparation[:20] + "..."
                    rich_text = form_rich_text(preparation)

                    if preparation_db_attribute_value:
                        recreate = self.should_recreate_attribute_assignment(
                            p,
                            preparation_db_attribute_value,
                        )

                        if recreate:
                            attribute_values_to_delete.append(
                                preparation_db_attribute_value
                            )
                            insert = True
                        else:
                            preparation_db_attribute_value.name = name
                            preparation_db_attribute_value.rich_text = rich_text
                            attribute_values_to_update.append(
                                preparation_db_attribute_value
                            )
                    else:
                        insert = True

                    if insert:
                        assigned_product_attribute_values_to_insert.append(
                            AttributeUtils.add_preparation_attribute_data(
                                product_id,
                                preparation,
                            )
                        )
                else:
                    if preparation_db_attribute_value:
                        attribute_values_to_delete.append(
                            preparation_db_attribute_value
                        )

                # DURATION AND DURATION UNIT

                duration_db_attribute_value = self.get_attribute_value(
                    product_id, "kdl-max_duration", duration_attribute_values
                )
                duration_unit_db_attribute_value = self.get_attribute_value(
                    product_id, "kdl-duration_unit", duration_unit_attribute_values
                )

                if data_product["duration"] and "N/A" not in str(
                    data_product["duration"]
                ):
                    insert_duration = False
                    insert_duration_unit = False

                    duration = int(data_product["duration"])

                    # DURATION
                    if duration_db_attribute_value:
                        recreate = self.should_recreate_attribute_assignment(
                            p,
                            duration_db_attribute_value,
                        )

                        if recreate:
                            attribute_values_to_delete.append(
                                duration_db_attribute_value
                            )
                            insert_duration = True
                        else:
                            duration_db_attribute_value.name = str(duration)
                            attribute_values_to_update.append(
                                duration_db_attribute_value
                            )
                    else:
                        insert_duration = True

                    if insert_duration:
                        assigned_product_attribute_values_to_insert.append(
                            AttributeUtils.add_numeric_attribute_data(
                                product_id,
                                duration,
                                AttributeUtils.attrubutes_ids["kdl-max_duration"],
                            )
                        )

                    # DURATION UNIT
                    if duration_unit_db_attribute_value:
                        recreate = self.should_recreate_attribute_assignment(
                            p,
                            duration_unit_db_attribute_value,
                        )

                        if recreate:
                            attribute_values_to_delete.append(
                                duration_unit_db_attribute_value
                            )
                            insert_duration_unit = True
                        else:
                            duration_unit_db_attribute_value.name = str(
                                KdlDurationUnitEnum.DAY.value
                            )
                            attribute_values_to_update.append(
                                duration_unit_db_attribute_value
                            )
                    else:
                        insert_duration_unit = True

                    if insert_duration_unit:
                        assigned_product_attribute_values_to_insert.append(
                            AttributeUtils.add_numeric_attribute_data(
                                product_id,
                                KdlDurationUnitEnum.DAY.value,
                                AttributeUtils.attrubutes_ids["kdl-duration_unit"],
                            )
                        )
                else:
                    if duration_db_attribute_value:
                        attribute_values_to_delete.append(duration_db_attribute_value)
                    if duration_unit_db_attribute_value:
                        attribute_values_to_delete.append(
                            duration_unit_db_attribute_value
                        )

                # BIOMARKERS

                if data_product["biomarkers"]:
                    biomarkers = [int(b) for b in data_product["biomarkers"]]

                    biomarkers_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "biomarkers"
                            ],
                            product_id=product_id,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        biomarkers_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product_id,
                            biomarkers,
                            biomarkers_attribute_values_ids,
                        )
                    )
                else:
                    biomarkers_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "biomarkers"
                            ],
                            product_id=product_id,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        biomarkers_assigned_product_attribute_values
                    )

                # MEDICAL EXAMS

                if data_product["medical_exams"]:
                    medical_exams = [int(m) for m in data_product["medical_exams"]]

                    medical_exams_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "medical_exams"
                            ],
                            product_id=product_id,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        medical_exams_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product_id,
                            medical_exams,
                            medical_exams_attribute_values_ids,
                        )
                    )
                else:
                    medical_exams_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "medical_exams"
                            ],
                            product_id=product_id,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        medical_exams_assigned_product_attribute_values
                    )

                # COLOR

                color_attribute_assigned_value = (
                    AssignedProductAttributeValue.objects.filter(
                        value__attribute_id=AttributeUtils.attrubutes_ids["color"],
                        product_id=product_id,
                    ).first()
                )

                if data_product["color_slug"]:
                    color_attribute_value_for_current_slug = color_attribute_values.get(
                        data_product["color_slug"]
                    )

                    if color_attribute_assigned_value:
                        if (
                            color_attribute_value_for_current_slug
                            and color_attribute_assigned_value.value
                            != color_attribute_value_for_current_slug
                        ):
                            color_attribute_assigned_value.value = (
                                color_attribute_value_for_current_slug
                            )
                            color_attribute_assigned_value.save()
                    elif color_attribute_value_for_current_slug:
                        assigned_product_attribute_values_to_insert.append(
                            AssignedProductAttributeValue(
                                value_id=color_attribute_value_for_current_slug.pk,
                                product_id=product_id,
                            )
                        )
                else:
                    if color_attribute_assigned_value:
                        color_attribute_assigned_value.delete()

                # POPULAR (FEARTURED CATEGORY & POPULAR COLLECTION)

                featured_attribute_assigned_value = (
                    AssignedProductAttributeValue.objects.filter(
                        value__attribute_id=AttributeUtils.attrubutes_ids["featured"],
                        product_id=product_id,
                    ).first()
                )

                is_popular = data_product["is_popular"]

                featured_true_slug = f"{AttributeUtils.attrubutes_ids['featured']}_true"
                featured_false_slug = (
                    f"{AttributeUtils.attrubutes_ids['featured']}_false"
                )

                featured_insert = False

                if featured_attribute_assigned_value:
                    compare_slug = (
                        featured_true_slug if is_popular == 1 else featured_false_slug
                    )
                    if featured_attribute_assigned_value.value.slug != compare_slug:
                        featured_insert = is_popular == 1
                        assigned_product_attribute_values_to_delete.append(
                            featured_attribute_assigned_value
                        )
                else:
                    featured_insert = is_popular == 1

                if featured_insert:
                    assigned_product_attribute_values_to_insert.append(
                        self.add_featured_attribute_data(
                            product_id, featured_attribute_values
                        )
                    )

                if popular_collection:
                    product_popular_collection_assignment = (
                        popular_collection_assignments.get(product_id)
                    )

                    if is_popular != 1 and product_popular_collection_assignment:
                        product_popular_collection_assignment.delete()
                    elif is_popular == 1 and not product_popular_collection_assignment:
                        collection_products_to_insert.append(
                            CollectionProduct(
                                collection_id=popular_collection.pk,
                                product_id=product_id,
                            )
                        )

        ProductVariantChannelListing.objects.filter(
            variant__sku__in=data_to_delist
        ).delete()

        AttributeValue.objects.filter(
            id__in=[av.pk for av in attribute_values_to_delete]
        ).delete()

        AssignedProductAttributeValue.objects.filter(
            id__in=[apav.pk for apav in assigned_product_attribute_values_to_delete]
        ).delete()

        Product.objects.bulk_update(
            products_to_update,
            ["category", "description", "description_plaintext"],
            batch_size=500,
        )
        AttributeValue.objects.bulk_update(
            attribute_values_to_update,
            ["name", "plain_text", "rich_text"],
            batch_size=500,
        )
        AssignedProductAttributeValue.objects.bulk_create(
            assigned_product_attribute_values_to_insert,
            batch_size=500,
        )
        CollectionProduct.objects.bulk_create(collection_products_to_insert)

        return
