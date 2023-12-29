import os

from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import DecimalValidator
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import exceptions as openpyxl_exceptions

from saleor.product.models import ProductVariantChannelListing, ProductVariant
from saleor.channel.models import Channel
from saleor.warehouse.models import Stock


class Command(BaseCommand):
    help = "Sync KDL channel product prices with XLSX file"

    def add_arguments(self, parser):
        parser.add_argument(
            "-f", "--filename", help="Path to XLSX file with target data"
        )
        parser.add_argument("-c", "--channel", help="Channel slug")

    def get_xls_header_row(self, sheet: Worksheet, header_first_col: str) -> int:
        header_row = [r for r in sheet["A"] if r.value and header_first_col in r.value]

        if not header_row:
            raise CommandError(
                f"No header has been found for the first column {header_first_col}. Please check source file!"
            )

        return header_row[0].row

    def price_validator(self, value):
        decimal_validator = DecimalValidator(
            max_digits=settings.DEFAULT_MAX_DIGITS,
            decimal_places=settings.DEFAULT_DECIMAL_PLACES,
        )

        dvalue = None
        try:
            dvalue = Decimal(value)
            decimal_validator(dvalue)
            error = None
        except InvalidOperation:
            error = f'Defined price "{value}" is not valid decimal value.'
        except ValidationError as e:
            error = e.messages[0]
        except Exception as e:
            error = str(e)

        return (value if dvalue is None else dvalue, error)

    def make_kdl_sku(self, sku: str) -> str:
        return f"KDL-{sku}"

    def handle(self, *args, **options):
        filename = options.get("filename")
        channel_slug = options.get("channel")

        if not filename or not os.path.isfile(filename):
            raise CommandError(f'XLSX file "{filename}" does not exist.')

        channel = Channel.objects.filter(slug=channel_slug).first()

        if not channel:
            raise CommandError(f'Selected channel "{channel}" is not valid!')

        try:
            wb = load_workbook(filename)
        except openpyxl_exceptions.InvalidFileException:
            raise CommandError(
                f"openpyxl does not support the old .xls file format, please use .xlsx file format."
            )

        sheet_1 = wb["Прейскурант"]
        sheet_2 = wb["Профили"]

        header1_row = self.get_xls_header_row(sheet_1, "Код услуги")
        header2_row = self.get_xls_header_row(sheet_2, "Код профиля")

        prices = {}
        prices_errors = []

        for value in sheet_1.iter_rows(min_row=header1_row + 1, values_only=True):
            price = value[7]
            if price:
                prices[self.make_kdl_sku(str(value[0]))] = price

        for value in sheet_2.iter_rows(min_row=header2_row + 1, values_only=True):
            price = value[8]
            if price:
                prices[self.make_kdl_sku(str(value[0]))] = price

        product_variant_channel_listings_q = (
            ProductVariantChannelListing.objects.filter(
                variant__sku__in=prices.keys(), channel=channel
            ).prefetch_related("channel", "variant")
        )

        updated_skus_prices = []
        inserted_skus_prices = []

        for pvcl in product_variant_channel_listings_q:
            sku = pvcl.variant.sku

            if sku:
                price = prices.pop(sku)

                if price:
                    validated_price, error = self.price_validator(price)

                    if error:
                        prices_errors.append(
                            {"sku": sku, "price": price, "error": error}
                        )
                        continue

                    if (
                        pvcl.price_amount == validated_price
                        and pvcl.discounted_price_amount == validated_price
                    ):
                        continue

                    updated_skus_prices.append(sku)

                    pvcl.price_amount = validated_price
                    pvcl.discounted_price_amount = validated_price

        variants = ProductVariant.objects.filter(sku__in=(prices.keys()))
        new_product_variant_channel_listings = []
        new_warehouse_stocks = []

        for variant in variants:
            if variant.sku:
                price = prices.get(variant.sku)
                inserted_skus_prices.append(variant.sku)

                new_product_variant_channel_listings.append(
                    ProductVariantChannelListing(
                        variant=variant,
                        channel=channel,
                        currency="RUB",
                        price_amount=price,
                        discounted_price_amount=price,
                    )
                )
                new_warehouse_stocks.append(
                    Stock(
                        warehouse_id="2e83f67b-a080-4710-9ee8-4bf3bc0e0b58",
                        product_variant=variant,
                        quantity=1,
                        quantity_allocated=0,
                    )
                )

        # update existing product variants listings
        ProductVariantChannelListing.objects.bulk_update(
            product_variant_channel_listings_q,
            ["price_amount", "discounted_price_amount"],
        )

        # create new product variants listings
        ProductVariantChannelListing.objects.bulk_create(
            new_product_variant_channel_listings
        )

        # create new stock for product variants listings
        Stock.objects.bulk_create(new_warehouse_stocks, ignore_conflicts=True)

        if prices_errors:
            print(prices_errors)
            return

        print("Updated SKUs prices:")
        print(updated_skus_prices)
        print("Inserted SKUs prices:")
        print(inserted_skus_prices)
        return
