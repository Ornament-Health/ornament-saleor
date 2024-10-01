from datetime import datetime
from itertools import chain
import os
import re

from django.core.management.base import BaseCommand, CommandError
from slugify import slugify
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import exceptions as openpyxl_exceptions

from saleor.attribute.models.base import AttributeValue
from saleor.attribute.models.product import AssignedProductAttributeValue
from saleor.ornament.vendors.attribute_utils import AttributeUtils
from saleor.ornament.vendors.utils import (
    fetch_medical_data,
    form_description,
)
from saleor.product.models import (
    Product,
    ProductVariant,
    ProductChannelListing,
    Category,
    ProductVariantChannelListing,
)
from saleor.warehouse.models import Stock
from saleor.channel.models import Channel


class Command(BaseCommand):
    help = "Import/Update GetTested products with XLSX file"
    known_header_rows = [
        "SKU Ornament",
        "Test",
        "Test method",
        "Category",
        "Subcategory",
        "GBP",
        "Description",
        "Biomarkers",
        "medical_exam_type_object",
        "is_hidden",
        "color slug",
        "human parts",
    ]
    vendor_name = "GetTested"
    vendor_currency = "GBP"
    vendor_warehouse_id = "08061547-682e-4196-bfcf-273a859aca25"

    def add_arguments(self, parser):
        parser.add_argument("filename", help="Path to XLSX file with target data")
        parser.add_argument(
            "--only_medical_data", action="store_true", help="Only update medical data"
        )

    def check_sheet_header_cols(self, sheet: Worksheet) -> None:
        sheet_header_cols: set = set(
            [s for s in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0]
        )

        if set(self.known_header_rows).difference(sheet_header_cols):
            raise CommandError(
                f"Sheet header does not contain all known columns: {self.known_header_rows}"
            )

    def check_row_required_data(self, row: tuple) -> bool:
        return all([row[0], row[1], row[2], row[3], row[4], row[5], row[6]])

    def add_test_method_attribute_data(
        self, product_id: int, test_method: str
    ) -> AssignedProductAttributeValue:
        attribute_id = AttributeUtils.attribute_ids["gettested_test-method"]
        name = test_method.replace("\n", ", ")
        slug = f"{product_id}_{attribute_id}"

        test_method_attribute_value = AttributeValue(
            name=name,
            attribute_id=attribute_id,
            slug=slug,
            plain_text=name,
            sort_order=attribute_id,
        )

        test_method_attribute_value.save()

        return AssignedProductAttributeValue(
            value_id=test_method_attribute_value.pk,
            product_id=product_id,
        )

    def get_medical_data_ids(self, col_data: str, ids: list[int]) -> list[int]:
        data_strings = col_data.split("\n")

        try:
            found_ids = chain.from_iterable(
                [re.findall(r"\d+", b.split("-")[-1]) for b in data_strings if b]
            )
            return [id for id in found_ids if int(id) in ids]
        except (IndexError, ValueError):
            raise CommandError(f"Can not parse medical data column: {col_data}")

    def get_data_from_worksheet(self, sheet: Worksheet) -> dict:
        medical_data = fetch_medical_data()

        return {
            s[0]: {
                "sku": s[0],
                "name": s[1],
                "test_method": s[2],
                "category": s[3],
                "subcategory": s[4],
                "price": s[5],
                "description": s[6],
                "biomarkers": self.get_medical_data_ids(
                    str(s[7]), medical_data.biomarker_ids
                ),
                "medical_exams": self.get_medical_data_ids(
                    str(s[8]), medical_data.medical_exams_ids
                ),
                "human_parts": self.get_medical_data_ids(
                    str(s[13]), medical_data.human_parts_ids
                ),
                "is_hidden": True if s[10] == 1 else False,
                "color_slug": s[11],
            }
            for s in sheet.iter_rows(min_row=2, values_only=True)
            if self.check_row_required_data(s)
        }

    def get_categories(self) -> dict[str, Category]:
        categories = Category.objects.all()
        return {c.name: c for c in categories}

    def get_medical_attribute_values_ids(self, attribute_name: str) -> dict[int, int]:
        medical_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attribute_ids[attribute_name]
        ).values_list("pk", "name")

        return {int(b[1]): b[0] for b in medical_attribute_values}

    def get_attribute_values(self, attribute_name: str) -> dict[str, AttributeValue]:
        attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attribute_ids[attribute_name]
        )
        return {v.slug: v for v in attribute_values}

    def handle(self, *args, **options):
        filename = options.get("filename")
        only_medical_data = options.get("only_medical_data")

        if not filename or not os.path.isfile(filename):
            raise CommandError(f'Source file "{filename}" does not exist.')

        try:
            wb = load_workbook(filename)
        except openpyxl_exceptions.InvalidFileException:
            raise CommandError(
                f"openpyxl does not support the old .xls file format, please use .xlsx file format."
            )

        sheet = wb["Import"]

        self.check_sheet_header_cols(sheet)

        data = self.get_data_from_worksheet(sheet)

        current_products = Product.objects.values_list("name", flat=True)

        categories = self.get_categories()

        gettested_channel = Channel.objects.filter(slug="uk-gettested").first()

        if not gettested_channel:
            raise CommandError(f"'uk-gettested' channel does not exist")

        data_to_update = {k: v for k, v in data.items() if k in current_products}
        data_to_insert = {
            k: v
            for k, v in data.items()
            if k not in current_products and not v["is_hidden"]
        }

        test_method_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attribute_ids["gettested_test-method"]
        )
        test_method_attribute_values = {v.slug: v for v in test_method_attribute_values}

        new_products = []

        assigned_product_attribute_values_to_insert: list[
            AssignedProductAttributeValue
        ] = []
        attribute_values_to_update: list[AttributeValue] = []
        assigned_product_attribute_values_to_delete: list[
            AssignedProductAttributeValue
        ] = []

        biomarkers_attribute_values_ids = self.get_medical_attribute_values_ids(
            "biomarkers"
        )
        medical_exams_attribute_values_ids = self.get_medical_attribute_values_ids(
            "medical_exams"
        )
        human_parts_attribute_values_ids = self.get_medical_attribute_values_ids(
            "human_parts"
        )

        if not only_medical_data:
            for d in data_to_insert.values():
                new_product_category = categories.get(d["subcategory"])

                if not new_product_category:
                    raise CommandError(f'Category not found: {d["subcategory"]}')

                new_products.append(
                    Product(
                        product_type_id=1,
                        name=d["sku"],
                        slug=slugify(d["sku"], lowercase=True, max_length=100),
                        description=form_description(d["name"], d["description"]),
                        description_plaintext=d["description"],
                        category=new_product_category,
                        search_index_dirty=True,
                    )
                )

            inserted_products = Product.objects.bulk_create(
                new_products, ignore_conflicts=False
            )

            product_variants = []
            product_channel_listings = []

            now = datetime.now()

            color_attribute_values = self.get_attribute_values("color")

            for product in inserted_products:
                product_variants.append(
                    ProductVariant(
                        sku=product.name,
                        name=self.vendor_name,
                        product=product,
                        track_inventory=False,
                    )
                )
                product_channel_listings.append(
                    ProductChannelListing(
                        channel=gettested_channel,
                        product=product,
                        visible_in_listings=True,
                        available_for_purchase_at=now,
                        currency=self.vendor_currency,
                        is_published=True,
                        published_at=now,
                    )
                )

                product_data = data_to_insert.get(product.name) or {}
                test_method = product_data.get("test_method")
                biomarkers = product_data.get("biomarkers")
                medical_exams = product_data.get("medical_exams")
                human_parts = product_data.get("human_parts")
                color_slug = product_data.get("color_slug")

                if test_method:
                    assigned_product_attribute_values_to_insert.append(
                        self.add_test_method_attribute_data(product.pk, test_method)
                    )

                if biomarkers:
                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            [int(b) for b in biomarkers],
                            biomarkers_attribute_values_ids,
                        )
                    )

                if medical_exams:
                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            [int(b) for b in medical_exams],
                            medical_exams_attribute_values_ids,
                        )
                    )

                if human_parts:
                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            [int(hp) for hp in human_parts],
                            human_parts_attribute_values_ids,
                        )
                    )

                if color_slug:
                    assigned_product_attribute_values_to_insert.append(
                        AttributeUtils.add_color_attribute_data(
                            product.pk,
                            color_slug,
                            color_attribute_values,
                        )
                    )

            inserted_variants = ProductVariant.objects.bulk_create(product_variants)
            ProductChannelListing.objects.bulk_create(product_channel_listings)

            new_product_variant_channel_listings = []
            new_warehouse_stocks = []

            for variant in inserted_variants:
                if variant.sku:
                    product_data = data_to_insert.get(variant.sku) or {}
                    price = product_data.get("price")

                    if price:
                        new_product_variant_channel_listings.append(
                            ProductVariantChannelListing(
                                variant=variant,
                                channel=gettested_channel,
                                currency=self.vendor_currency,
                                price_amount=price,
                                discounted_price_amount=price,
                            )
                        )
                        new_warehouse_stocks.append(
                            Stock(
                                warehouse_id=self.vendor_warehouse_id,
                                product_variant=variant,
                                quantity=1,
                                quantity_allocated=0,
                            )
                        )

            # create new product variants listings
            ProductVariantChannelListing.objects.bulk_create(
                new_product_variant_channel_listings
            )

            # create new stock for product variants listings
            Stock.objects.bulk_create(new_warehouse_stocks)

        # UPDATE
        products_to_update = Product.objects.filter(name__in=data_to_update.keys())
        product_variant_channel_listings_to_update = (
            ProductVariantChannelListing.objects.filter(
                variant__sku__in=data_to_update.keys(), channel=gettested_channel
            ).prefetch_related("channel", "variant")
        )

        data_to_delist = [k for k, v in data_to_update.items() if v["is_hidden"] == 1]

        for product in products_to_update:
            data_product = data_to_update.get(product.name)

            if data_product:
                sku = data_product.get("sku")
                name = data_product.get("name")
                description = data_product.get("description")
                test_method = data_product.get("test_method")
                biomarkers = data_product.get("biomarkers")
                medical_exams = data_product.get("medical_exams")
                human_parts = data_product.get("human_parts")
                subcategory = data_product["subcategory"]
                color_slug = data_product["color_slug"]

                product_category = categories.get(subcategory)

                if not product_category:
                    raise CommandError(f"Category not found: {subcategory}")

                if not only_medical_data:
                    product.category = product_category

                    if sku:
                        product.name = sku
                        product.search_index_dirty = True

                    if name and description:
                        product.description = form_description(name, description)
                        product.description_plaintext = description
                        product.search_index_dirty = True

                    if test_method:
                        slug = f"{product.pk}_{AttributeUtils.attribute_ids['gettested_test-method']}"
                        db_attribute_value = test_method_attribute_values.get(slug)

                        if db_attribute_value:
                            db_attribute_value.name = test_method
                            db_attribute_value.plain_text = name
                            attribute_values_to_update.append(db_attribute_value)
                        else:
                            assigned_product_attribute_values_to_insert.append(
                                self.add_test_method_attribute_data(
                                    product.pk, test_method
                                )
                            )

                    color_attribute_assigned_value = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attribute_ids["color"],
                            product_id=product.pk,
                        ).first()
                    )

                    if color_slug:
                        color_attribute_value_for_current_slug = (
                            color_attribute_values.get(color_slug)
                        )

                        if color_attribute_assigned_value:
                            if (
                                color_attribute_value_for_current_slug
                                and color_attribute_assigned_value.value
                                != color_attribute_value_for_current_slug
                            ):
                                color_attribute_assigned_value.value = (
                                    color_attribute_value_for_current_slug
                                )
                                color_attribute_assigned_value.save()
                        elif color_attribute_value_for_current_slug:
                            assigned_product_attribute_values_to_insert.append(
                                AssignedProductAttributeValue(
                                    value_id=color_attribute_value_for_current_slug.pk,
                                    product_id=product.pk,
                                )
                            )
                    else:
                        if color_attribute_assigned_value:
                            color_attribute_assigned_value.delete()

                if biomarkers:
                    biomarkers = [int(b) for b in biomarkers]

                    biomarkers_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attribute_ids[
                                "biomarkers"
                            ],
                            product_id=product.pk,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        biomarkers_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            biomarkers,
                            biomarkers_attribute_values_ids,
                        )
                    )
                else:
                    biomarkers_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attribute_ids[
                                "biomarkers"
                            ],
                            product_id=product.pk,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        biomarkers_assigned_product_attribute_values
                    )

                if medical_exams:
                    medical_exams = [int(b) for b in medical_exams]

                    medical_exams_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attribute_ids[
                                "medical_exams"
                            ],
                            product_id=product.pk,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        medical_exams_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            medical_exams,
                            medical_exams_attribute_values_ids,
                        )
                    )
                else:
                    medical_exams_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attribute_ids[
                                "medical_exams"
                            ],
                            product_id=product.pk,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        medical_exams_assigned_product_attribute_values
                    )

                if human_parts:
                    human_parts = [int(hp) for hp in human_parts]

                    human_parts_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attribute_ids[
                                "human_parts"
                            ],
                            product_id=product.pk,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        human_parts_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            human_parts,
                            human_parts_attribute_values_ids,
                        )
                    )
                else:
                    human_parts_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attribute_ids[
                                "human_parts"
                            ],
                            product_id=product.pk,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        human_parts_assigned_product_attribute_values
                    )

        if not only_medical_data:
            for (
                product_variant_channel_listing
            ) in product_variant_channel_listings_to_update:
                sku = product_variant_channel_listing.variant.sku

                if sku:
                    data_product = data_to_update.get(sku) or {}
                    price = data_product.get("price")

                    if price:
                        product_variant_channel_listing.price_amount = price
                        product_variant_channel_listing.discounted_price_amount = price

        # TODO: https://github.com/Ornament-Health/ornament-saleor/pull/7#discussion_r1447222450
        AssignedProductAttributeValue.objects.filter(
            id__in=[apav.pk for apav in assigned_product_attribute_values_to_delete]
        ).delete()

        AssignedProductAttributeValue.objects.bulk_create(
            assigned_product_attribute_values_to_insert,
            batch_size=500,
        )

        AttributeValue.objects.bulk_update(
            attribute_values_to_update, ["name", "plain_text"]
        )

        Product.objects.bulk_update(
            products_to_update,
            ["name", "description", "search_index_dirty"],
        )

        ProductVariantChannelListing.objects.bulk_update(
            product_variant_channel_listings_to_update,
            ["price_amount", "discounted_price_amount"],
        )

        ProductVariantChannelListing.objects.filter(
            variant__sku__in=data_to_delist
        ).delete()

        return
