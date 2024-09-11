from datetime import datetime
import json
import os
from typing import Optional
from django.conf import settings

from django.core.management.base import BaseCommand, CommandError
from slugify import slugify
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import exceptions as openpyxl_exceptions

from saleor.attribute.models.base import AttributeValue
from saleor.attribute.models.product import AssignedProductAttributeValue
from saleor.ornament.vendors.attribute_utils import AttributeUtils
from saleor.ornament.vendors.utils import (
    form_description,
)
from saleor.product.models import (
    Product,
    ProductVariant,
    ProductChannelListing,
    Category,
    ProductVariantChannelListing,
)
from saleor.channel.models import Channel
from saleor.warehouse.models import Stock


class Command(BaseCommand):
    help = "Import/Update DarDoc products with XLSX file (using KDL SKU database)"
    known_header_cols = [
        "SKU KDL",
        "Test Name",
        "AED PRICE",
        "TAT Days",
    ]
    vendor_name = "DarDoc"
    vendor_channel_slug = "uae-dardoc"
    vendor_currency = "AED"
    vendor_warehouse_id = "ca8c74be-3d0a-4e52-a8cb-74b3f23f4b3b"
    vendor_default_product_type_id = 1
    kdl_products_data_path = os.path.join(
        settings.PROJECT_ROOT, "saleor", "ornament", "vendors", "dardoc", "data"
    )

    def add_arguments(self, parser):
        parser.add_argument("filename", help="Path to XLSX file with target data")

    def check_sheet_header_cols(self, sheet: Worksheet) -> None:
        sheet_header_cols: set = set(
            [s for s in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0]
        )

        if set(self.known_header_cols).difference(sheet_header_cols):
            raise CommandError(
                f"Sheet header does not contain all known columns: {self.known_header_cols}"
            )

    def check_row_valid(self, row: tuple) -> bool:
        return (
            all([row[0], row[4], row[5], row[6]])
            and isinstance(row[0], str)
            and row[0] != "none"
        )

    def get_kdl_sku_data(self, row: tuple, kdl_products_data: dict) -> Optional[dict]:
        possible_sku = row[0].split("\n")[0]
        kdl_data = kdl_products_data.get(possible_sku)

        if not kdl_data:
            return None

        return {"sku": possible_sku, "kdl_data": kdl_data}

    def get_attribute_value(
        self,
        product_id: int,
        attribute_name: str,
        known_values: dict[str, AttributeValue],
    ) -> Optional[AttributeValue]:
        slug = f"{product_id}_{AttributeUtils.attrubutes_ids[attribute_name]}"
        return known_values.get(slug)

    def should_recreate_attribute_assignment(
        self,
        product: Product,
        db_attribute_value: AttributeValue,
    ) -> bool:
        assigned_value = product.attributevalues.filter(
            value_id=db_attribute_value.pk
        ).first()

        return not assigned_value

    def get_kdl_products_data(self) -> dict:
        path = os.path.join(self.kdl_products_data_path, "kdl_products_data.json")
        with open(path, "r") as kdl_products_data_file:
            return json.load(kdl_products_data_file)

    def get_medical_attributes_ids(self, atrribute: str) -> dict[int, int]:
        biomarkers_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids[atrribute]
        ).values_list("pk", "name")
        return {int(b[1]): b[0] for b in biomarkers_attribute_values}

    def handle(self, *args, **options):
        filename = options.get("filename")

        if not filename or not os.path.isfile(filename):
            raise CommandError(f'Source file "{filename}" does not exist.')

        try:
            wb = load_workbook(filename)
        except openpyxl_exceptions.InvalidFileException:
            raise CommandError(
                f"openpyxl does not support the old .xls file format, please use .xlsx file format."
            )

        kdl_products_data = self.get_kdl_products_data()

        sheet = wb["Import"]

        self.check_sheet_header_cols(sheet)

        data = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not self.check_row_valid(row):
                continue

            kdl_product = self.get_kdl_sku_data(row, kdl_products_data)

            if not kdl_product:
                print(f"Row not found in KDL SKU database: {row}")
                continue

            sku = kdl_product["sku"]
            kdl_data = kdl_product["kdl_data"]

            data[sku] = {
                "sku": sku,
                "name": row[4],
                "duration": row[6],
                "price": row[5],
                "biomarkers": kdl_data["biomarkers"],
                "medical_exams": kdl_data["medical_exams"],
            }

        current_products = Product.objects.values_list("name", flat=True)

        dardoc_category = Category.objects.filter(slug="dardoc").first()
        dardoc_channel = Channel.objects.filter(slug=self.vendor_channel_slug).first()

        if not dardoc_category or not dardoc_channel:
            raise CommandError(
                f"No `dardoc` category or `{self.vendor_channel_slug}` channel has been found!"
            )

        data_to_update = {k: v for k, v in data.items() if k in current_products}
        data_to_insert = {k: v for k, v in data.items() if k not in current_products}

        dardoc_duration_attribute_values = AttributeValue.objects.filter(
            attribute_id=AttributeUtils.attrubutes_ids["dardoc-duration"]
        )
        dardoc_duration_attribute_values = {
            v.slug: v for v in dardoc_duration_attribute_values
        }

        new_products = []

        for d in data_to_insert.values():
            new_products.append(
                Product(
                    product_type_id=self.vendor_default_product_type_id,
                    name=d["sku"],
                    slug=slugify(d["sku"], lowercase=True, max_length=100),
                    description=form_description(d["name"], ""),
                    description_plaintext="",
                    category_id=dardoc_category.pk,
                    search_index_dirty=True,
                )
            )

        inserted_products = Product.objects.bulk_create(
            new_products, ignore_conflicts=False
        )

        product_variants = []
        product_channel_listings = []
        assigned_product_attribute_values = []

        now = datetime.now()

        biomarkers_attribute_values_ids = self.get_medical_attributes_ids("biomarkers")
        medical_exams_attribute_values_ids = self.get_medical_attributes_ids(
            "medical_exams"
        )

        for product in inserted_products:
            data_product = data.get(product.name)

            if data_product:
                product_variants.append(
                    ProductVariant(
                        sku=f'{self.vendor_name}-{data_product["sku"]}',
                        name=self.vendor_name,
                        product=product,
                        track_inventory=False,
                    )
                )

                product_channel_listings.append(
                    ProductChannelListing(
                        channel=dardoc_channel,
                        product=product,
                        visible_in_listings=True,
                        available_for_purchase_at=now,
                        currency=self.vendor_currency,
                        is_published=True,
                        published_at=now,
                    )
                )

                if data_product["duration"]:
                    assigned_product_attribute_values.append(
                        AttributeUtils.add_numeric_attribute_data(
                            product.pk,
                            int(data_product["duration"]),
                            AttributeUtils.attrubutes_ids["dardoc-duration"],
                        )
                    )

                if data_product["biomarkers"]:
                    assigned_product_attribute_values += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            [int(b) for b in data_product["biomarkers"]],
                            biomarkers_attribute_values_ids,
                        )
                    )

                if data_product["medical_exams"]:
                    assigned_product_attribute_values += (
                        AttributeUtils.add_medical_attributes_data(
                            product.pk,
                            [int(m) for m in data_product["medical_exams"]],
                            medical_exams_attribute_values_ids,
                        )
                    )

        ProductChannelListing.objects.bulk_create(product_channel_listings)
        inserted_variants = ProductVariant.objects.bulk_create(product_variants)

        AssignedProductAttributeValue.objects.bulk_create(
            assigned_product_attribute_values
        )

        new_product_variant_channel_listings = []
        new_warehouse_stocks = []

        for variant in inserted_variants:
            if variant.sku:
                product_data = data_to_insert.get(variant.product.name) or {}
                price = product_data.get("price")

                if price:
                    new_product_variant_channel_listings.append(
                        ProductVariantChannelListing(
                            variant=variant,
                            channel=dardoc_channel,
                            currency=self.vendor_currency,
                            price_amount=price,
                            discounted_price_amount=price,
                        )
                    )
                    new_warehouse_stocks.append(
                        Stock(
                            warehouse_id=self.vendor_warehouse_id,
                            product_variant=variant,
                            quantity=1,
                            quantity_allocated=0,
                        )
                    )

        # create new product variants listings
        ProductVariantChannelListing.objects.bulk_create(
            new_product_variant_channel_listings
        )

        # create new stock for product variants listings
        Stock.objects.bulk_create(new_warehouse_stocks)

        # UPDATE
        data_to_delist = []

        products_to_update = Product.objects.filter(
            name__in=data_to_update.keys()
        ).prefetch_related("attributevalues")

        product_variant_channel_listings_to_update = (
            ProductVariantChannelListing.objects.filter(
                variant__sku__in=data_to_update.keys(), channel=dardoc_channel
            ).prefetch_related("channel", "variant")
        )

        attribute_values_to_update: list[AttributeValue] = []
        assigned_product_attribute_values_to_insert: list[
            AssignedProductAttributeValue
        ] = []
        attribute_values_to_delete: list[AttributeValue] = []
        assigned_product_attribute_values_to_delete: list[
            AssignedProductAttributeValue
        ] = []

        missed_product_channel_listings = []
        missed_product_variants = []
        missed_product_variants_channel_listings = []
        missed_warehouse_stocks = []

        for p in products_to_update:

            if dardoc_channel not in [p.channel for p in p.channel_listings.all()]:
                missed_product_channel_listings.append(
                    ProductChannelListing(
                        channel=dardoc_channel,
                        product=p,
                        visible_in_listings=True,
                        available_for_purchase_at=now,
                        currency=self.vendor_currency,
                        is_published=True,
                        published_at=now,
                    )
                )

            if self.vendor_name not in [v.name for v in p.variants.all()]:
                missed_product_variants.append(
                    ProductVariant(
                        sku=f"{self.vendor_name}-{p.name}",
                        name=self.vendor_name,
                        product=p,
                        track_inventory=False,
                    )
                )

            data_product = data_to_update.get(p.name)
            product_id = p.pk

            if data_product:
                p.category = dardoc_category

                # TODO DESCRIPTION!
                # p.description = form_description(data_product["name"], "")
                # p.description_plaintext = ""

                # DURATION

                duration_db_attribute_value = self.get_attribute_value(
                    product_id, "dardoc-duration", dardoc_duration_attribute_values
                )

                if data_product["duration"]:
                    insert_duration = False

                    duration = int(data_product["duration"])

                    if duration_db_attribute_value:
                        recreate = self.should_recreate_attribute_assignment(
                            p,
                            duration_db_attribute_value,
                        )

                        if recreate:
                            attribute_values_to_delete.append(
                                duration_db_attribute_value
                            )
                            insert_duration = True
                        else:
                            duration_db_attribute_value.name = str(duration)
                            attribute_values_to_update.append(
                                duration_db_attribute_value
                            )
                    else:
                        insert_duration = True

                    if insert_duration:
                        assigned_product_attribute_values_to_insert.append(
                            AttributeUtils.add_numeric_attribute_data(
                                product_id,
                                duration,
                                AttributeUtils.attrubutes_ids["dardoc-duration"],
                            )
                        )

                # BIOMARKERS

                if data_product["biomarkers"]:
                    biomarkers = [int(b) for b in data_product["biomarkers"]]

                    biomarkers_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "biomarkers"
                            ],
                            product_id=product_id,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        biomarkers_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product_id,
                            biomarkers,
                            biomarkers_attribute_values_ids,
                        )
                    )
                else:
                    biomarkers_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "biomarkers"
                            ],
                            product_id=product_id,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        biomarkers_assigned_product_attribute_values
                    )

                # MEDICAL EXAMS

                if data_product["medical_exams"]:
                    medical_exams = [int(m) for m in data_product["medical_exams"]]

                    medical_exams_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "medical_exams"
                            ],
                            product_id=product_id,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        medical_exams_assigned_product_attribute_values
                    )

                    assigned_product_attribute_values_to_insert += (
                        AttributeUtils.add_medical_attributes_data(
                            product_id,
                            medical_exams,
                            medical_exams_attribute_values_ids,
                        )
                    )
                else:
                    medical_exams_assigned_product_attribute_values = (
                        AssignedProductAttributeValue.objects.filter(
                            value__attribute_id=AttributeUtils.attrubutes_ids[
                                "medical_exams"
                            ],
                            product_id=product_id,
                        )
                    )
                    assigned_product_attribute_values_to_delete += (
                        medical_exams_assigned_product_attribute_values
                    )

        ProductChannelListing.objects.bulk_create(missed_product_channel_listings)
        inserted_missed_variants = ProductVariant.objects.bulk_create(
            missed_product_variants
        )

        for missed_variant in inserted_missed_variants:
            if missed_variant.sku:
                product_data = data_to_update.get(missed_variant.product.name) or {}
                price = product_data.get("price")

                if price:
                    missed_product_variants_channel_listings.append(
                        ProductVariantChannelListing(
                            variant=missed_variant,
                            channel=dardoc_channel,
                            currency=self.vendor_currency,
                            price_amount=price,
                            discounted_price_amount=price,
                        )
                    )
                    missed_warehouse_stocks.append(
                        Stock(
                            warehouse_id=self.vendor_warehouse_id,
                            product_variant=missed_variant,
                            quantity=1,
                            quantity_allocated=0,
                        )
                    )

        # create new product variants listings
        ProductVariantChannelListing.objects.bulk_create(
            missed_product_variants_channel_listings
        )

        # create new stock for product variants listings
        Stock.objects.bulk_create(missed_warehouse_stocks)

        for (
            product_variant_channel_listing
        ) in product_variant_channel_listings_to_update:
            sku = product_variant_channel_listing.variant.sku

            if sku:
                data_product = data_to_update.get(sku) or {}
                price = data_product.get("price")

                if price:
                    product_variant_channel_listing.price_amount = price
                    product_variant_channel_listing.discounted_price_amount = price
                else:
                    data_to_delist.append(sku)

        AttributeValue.objects.filter(
            id__in=[av.pk for av in attribute_values_to_delete]
        ).delete()

        AssignedProductAttributeValue.objects.filter(
            id__in=[apav.pk for apav in assigned_product_attribute_values_to_delete]
        ).delete()

        Product.objects.bulk_update(
            products_to_update,
            ["category", "description", "description_plaintext"],
            batch_size=500,
        )
        AttributeValue.objects.bulk_update(
            attribute_values_to_update,
            ["name", "plain_text", "rich_text"],
            batch_size=500,
        )
        AssignedProductAttributeValue.objects.bulk_create(
            assigned_product_attribute_values_to_insert,
            batch_size=500,
        )
        ProductVariantChannelListing.objects.bulk_update(
            product_variant_channel_listings_to_update,
            ["price_amount", "discounted_price_amount"],
        )
        ProductVariantChannelListing.objects.filter(
            variant__sku__in=data_to_delist
        ).delete()

        return
